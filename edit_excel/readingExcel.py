import openpyxl
import os

os.chdir('/home/lenny/escuela')

workbook = openpyxl.load_workbook('Calificaciones.xlsx')

print(type(workbook))

print(type(workbook['PROM GENERAL'])) #SHEET 1

print(workbook.sheetnames) #Get all the names of the sheets in the workbook

sheet = workbook['PROM GENERAL']
cell = sheet['C12']

print(cell.value) #It is also valid to write sheet['C12'].values or going even deeper.

#It is also valid to...
print(sheet.cell(row=12, column=3)) #Useful especially with for loops