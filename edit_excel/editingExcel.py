import openpyxl
import os

os.chdir('/home/lenny/development/learning/python/edit_excel/excelFiles')

wb = openpyxl.Workbook() #It is a brand new wb, it hasnt created the file tho.

sheet = wb['Sheet']
cell_a1 = sheet['A1'] #Currently the value is None

#To asign a value to a the cells, is as simple as...
#NOTE The asign has to be done in the notation sheet['...'], otherwise, python will read a var as 'cell_a1' as a overwrited variable.

sheet["A1"] = 90

wb.save('example.xlsx') #Now it's saved. Yeiii!

#To create a new Sheet go for...

sheet2 = wb.create_sheet() #You can aldo modify the var (index=... , title=...)

sheet2.title = 'Test uwu'

#Note that you can create diferent versions of the same wb depending at what lenght ur gonna save it

wb.save('example2.xlsx')

print(wb.sheetnames)